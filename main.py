import pandas as pd
import plotly.express as px

from jinja2 import Template
from openpyxl import load_workbook


stages = ['Data Generated', 'Data Formatted', 'Data Staged', 'Protocols.io']
d = {
    'Working Groups': [],
    'Data Generated': [],
    'Data Formatted': [],
    'Data Staged': [],
    'Protocols.io': []
}


def get_data(xlsx_file: str):
    """
    Takes an excel file and returns a dictionary of working groups and
    an enumerated count of the amount of data they have at what stage.
    """
    wb = load_workbook(xlsx_file)
    ws = wb.active

    data_generated = 0
    data_formatted = 0
    data_staged = 0
    protocols = 0

    for row in ws.iter_rows(min_row=7, max_col=12, max_row=28, values_only=True):
        if row[1] and row[1].startswith('ADGC '):
            # new working group seen; all following rows underneath are associated with this group
            current_group = row[1].split('-')[1].strip()
            d['Working Groups'].append(current_group)

            # reset all of the data from the previous group
            d['Data Generated'].append(data_generated)
            data_generated = 0
            d['Data Formatted'].append(data_formatted)
            data_formatted = 0
            d['Data Staged'].append(data_staged)
            data_staged = 0
            d['Protocols.io'].append(protocols)
            protocols = 0
        else:
            # rows under their working groups are parsed for the stage of progress they are at
            if row[2] != 'Not Yet Started':
                data_generated += 1
            if row[6] != 'Not Yet Started':
                data_formatted += 1
                data_staged += 1
            if row[8] != 'Not Yet Started':
                protocols += 1

    return pd.DataFrame(d)


def generate_figure(df):
    """Takes a pandas dataframe and returns a plotly bargraph figure object."""
    fig = px.bar(
        data_frame=df,
        x='Working Groups',
        y=stages,
        text=[stages for i in range(4)], text_auto=True,
        opacity=0.9,
        orientation='v',
        barmode='group',
        title='UCSC Data Submission Tracker',
        labels={
            'value': 'Files Submitted',
            'variable': 'Processing Stage'
        }
    )
    fig.update_yaxes(range=[0, 20])
    fig.update_traces(textposition='outside')
    fig.update_layout(legend=dict(
        orientation="h",
        entrywidth=100,
        yanchor="bottom",
        y=1.02,
        xanchor="right",
        x=1
    ))
    return fig


def generate_website(fig):
    """Renders a website from a template with a plotly figure named 'fig' embedded in it."""
    input_template_path = 'template.html'
    output_html_path = 'index.html'

    plotly_jinja_data = {'fig': fig.to_html(full_html=False)}

    with open(input_template_path) as f:
        template = Template(f.read())

    with open(output_html_path, 'w') as f:
        f.write(template.render(plotly_jinja_data))

    print('Success!')


def main():
    data = get_data(xlsx_file='data/SSPsyGene_Data_Tracker.xlsx')
    fig = generate_figure(data)
    generate_website(fig)


if __name__ == '__main__':
    main()
