import pandas as pd
import plotly.express as px

from jinja2 import Template
from openpyxl import load_workbook


d = {
    'Working Groups': [],
    'Expected Datasets': [],
    'Data Staging Begun': [],
    'Data Staged': [],
    'Protocols.io Begun': [],
    'Protocols.io Completed': []
}
stages = [k for k in list(d.keys())[1:]]
print(stages)


def get_data(xlsx_file: str):
    """
    Takes an excel file and returns a dictionary of working groups and
    an enumerated count of the amount of data they have at what stage.
    """
    wb = load_workbook(xlsx_file)
    ws = wb.active

    not_yet_begun = 0
    data_staging_begun = 0
    data_staged = 0
    protocols_begun = 0
    protocols_completed = 0

    for row in ws.iter_rows(min_row=7, max_col=12, max_row=100, values_only=True):
        if row[1] and row[1].startswith('ADGC ') and len(d['Working Groups']) != len(d['Expected Datasets']):
            # reset all of the data from the previous group
            d['Expected Datasets'].append(not_yet_begun)
            not_yet_begun = 0
            d['Data Staging Begun'].append(data_staging_begun)
            data_staging_begun = 0
            d['Data Staged'].append(data_staged)
            data_staged = 0
            d['Protocols.io Begun'].append(protocols_begun)
            protocols_begun = 0
            d['Protocols.io Completed'].append(protocols_completed)
            protocols_completed = 0
            print(current_group)
        if row[1] and row[1].startswith('ADGC '):
            # new working group seen; all following rows underneath are associated with this group
            current_group = row[1].split('-')[1].strip()
            d['Working Groups'].append(current_group)
        else:
            print(row)
            # rows under their working groups are parsed for the stage of progress they are at
            if row[3] == 'Not Yet Started':
                not_yet_begun += 1
            elif row[3] == 'Submitted':
                data_staged += 1
            elif row[3]:
                data_staging_begun += 1

            if row[10] and row[10] == 'Submitted':
                protocols_completed += 1
            elif row[10] and row[10] != 'Not Yet Started':
                protocols_begun += 1
    d['Expected Datasets'].append(not_yet_begun)
    d['Data Staging Begun'].append(data_staging_begun)
    d['Data Staged'].append(data_staged)
    d['Protocols.io Begun'].append(protocols_begun)
    d['Protocols.io Completed'].append(protocols_completed)
    return pd.DataFrame(d)


def generate_figure(df):
    """Takes a pandas dataframe and returns a plotly bargraph figure object."""
    fig = px.bar(
        data_frame=df,
        x='Working Groups',
        y=stages,
        text=[stages for i in range(len(df))], text_auto=True,
        opacity=0.9,
        orientation='v',
        barmode='group',
        title='UCSC Data Submission Tracker',
        labels={
            'value': 'Datasets',
            'variable': 'Processing Stage'
        }
    )
    fig.update_yaxes(range=[0, 20])
    fig.update_traces(textposition='outside')
    fig.update_layout(legend=dict(
        orientation="h",
        entrywidth=140,
        yanchor="bottom",
        y=1.02,
        xanchor="right",
        x=1
    ))
    return fig


def generate_website(fig):
    """Renders a website from a template with a plotly figure named 'fig' embedded in it."""
    input_template_path = 'template.html'
    output_html_path = 'index.html'

    plotly_jinja_data = {'fig': fig.to_html(full_html=False)}

    with open(input_template_path) as f:
        template = Template(f.read())

    with open(output_html_path, 'w') as f:
        f.write(template.render(plotly_jinja_data))

    print('Success!')


def main():
    data = get_data(xlsx_file='data/SSPsyGene_Data_Tracker_v2.xlsx')
    fig = generate_figure(data)
    generate_website(fig)


if __name__ == '__main__':
    main()
